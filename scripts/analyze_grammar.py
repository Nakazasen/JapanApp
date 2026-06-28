import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))


import openpyxl
import os

file_path = r"C:\ProgramData\Sandbox\Projects\EnglishApp\[Tailieutiengnhat.net]_tu-vung-tieng-nhat-n1-day-du.xlsx"
sheet_name = "Êñ∞ÂÆåÂÅEÊñÅE≥ïN1"

def analyze():
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print(f"Loading {sheet_name}...")
    # Use read_only to be faster
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet not found. Available: {wb.sheetnames}")
        return
        
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
    
    print("Headers:")
    print(rows[0])
    
    print("\nSample Data:")
    for row in rows[1:]:
        print(row)

if __name__ == "__main__":
    analyze()

