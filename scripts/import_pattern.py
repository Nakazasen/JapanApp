import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))


import openpyxl
from sqlmodel import Session, create_engine, select
from frontend.core.config import settings
from frontend.models.vocab import JpVocabItem, VocabTopic
import os

# --- Configurations ---
project_root = r"C:\ProgramData\Sandbox\Projects\EnglishApp"
file_path = os.path.join(project_root, "[Tailieutiengnhat.net]_tu-vung-tieng-nhat-n1-day-du.xlsx")
sheet_name = "パターン語彙N1"
SOURCE = "Pattern Goi N1"
LEVEL = "N1"
USER_ID = 1 # Admin user
TOPIC_NAME = "Pattern Goi N1 (Niềm tin)" # Names like this help distinguish from Mimikara

def import_data():
    database_url = f"sqlite:///{settings.db_path}"
    engine = create_engine(database_url)
    
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    print(f"Opening Excel sheet: {sheet_name}...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found.")
        return
        
    ws = wb[sheet_name]
    
    vocab_list = []
    current_main_word = None
    
    print("Processing rows...")
    # Row 1 is header: No, Từ vựng, Phiên âm, Hán Việt, ÁEnghĩa
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        no = str(row[0]) if row[0] is not None else ""
        kanji = row[1]
        kana = row[2]
        han_viet = row[3]
        meaning = row[4]
        
        if not kanji:
            continue
            
        # Check if it's a main number (e.g. "1", "2", "3")
        is_main = False
        try:
            if no and '.' not in no:
                # If it's just a number like 1, 2, 3
                num = float(no)
                if num.is_integer():
                    is_main = True
        except ValueError:
            pass
            
        if is_main:
            # Create a new main word entry
            current_main_word = {
                "user_id": USER_ID,
                "word_kanji": str(kanji),
                "word_kana": str(kana) if kana else "",
                "han_viet": str(han_viet) if han_viet else "",
                "meaning_vi": str(meaning) if meaning else "",
                "example_jp": "",
                "level": LEVEL,
                "source_material": SOURCE,
                "mastery_status": "new"
            }
            vocab_list.append(current_main_word)
        else:
            # It's a sub-item (1.1, 1.2...), add as example or its own entry?
            # User usually wants them all. Mimikara script added them as examples.
            # But in Pattern Goi, 1.1 is often a DIFFERENT word with same reading.
            # Let's add them as separate words if they have Kanji.
            item = {
                "user_id": USER_ID,
                "word_kanji": str(kanji),
                "word_kana": str(kana) if kana else "",
                "han_viet": str(han_viet) if han_viet else "",
                "meaning_vi": str(meaning) if meaning else "",
                "example_jp": "",
                "level": LEVEL,
                "source_material": SOURCE,
                "mastery_status": "new"
            }
            vocab_list.append(item)

    print(f"Collected {len(vocab_list)} entries.")
    
    # Save to Database
    with Session(engine) as session:
        # 1. Ensure Topic exists
        statement = select(VocabTopic).where(VocabTopic.name == TOPIC_NAME)
        topic = session.exec(statement).first()
        if not topic:
            topic = VocabTopic(
                name=TOPIC_NAME,
                description="Từ vựng Pattern Goi N1 nhập từ Excel",
                user_id=USER_ID,
                lang="jp",
                icon="📖"
            )
            session.add(topic)
            session.commit()
            session.refresh(topic)
        
        count = 0
        for item_data in vocab_list:
            item_data["topic_id"] = topic.id
            
            # Avoid direct duplicates if possible (simple check)
            statement = select(JpVocabItem).where(
                JpVocabItem.word_kanji == item_data["word_kanji"],
                JpVocabItem.word_kana == item_data["word_kana"],
                JpVocabItem.source_material == SOURCE
            )
            existing = session.exec(statement).first()
            
            if not existing:
                db_item = JpVocabItem(**item_data)
                session.add(db_item)
                count += 1
        
        session.commit()
        print(f"Successfully imported {count} new entries to database.")

if __name__ == "__main__":
    import_data()

