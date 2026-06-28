import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))


import openpyxl
from openpyxl.cell.cell import MergedCell

file_path = r"C:\Users\Admin\Downloads\[Tailieutiengnhat.net]_tu-vung-tieng-nhat-n1-day-du.xlsx"
sheet_name = "パターン語彙N1"

print(f"Opening workbook sheet: {sheet_name}...")
wb = openpyxl.load_workbook(file_path)
if sheet_name not in wb.sheetnames:
    print(f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    exit(1)

ws = wb[sheet_name]

print("Re-indexing column 1...")
current_base = None
sub_index = 0
count = 0

# Skip header (Row 1), start from Row 2
for row in ws.iter_rows(min_row=2):
    index_cell = row[0]
    
    # Handle MergedCells if any
    if isinstance(index_cell, MergedCell):
        continue

    val = index_cell.value
    
    # Check if this is a main number (int or float that represents an integer)
    is_main = False
    try:
        if val is not None:
            # Clean numeric string if needed
            num_str = str(val).split('.')[0] if '.' in str(val) and str(val).split('.')[1] == '0' else str(val)
            num = float(val)
            if num.is_integer():
                current_base = int(num)
                sub_index = 0
                is_main = True
    except (ValueError, TypeError):
        pass

    if not is_main and current_base is not None:
        # Check if the row has content in column 2 (Vocabulary)
        vocab_val = row[1].value
        if vocab_val is not None:
            sub_index += 1
            new_val = f"{current_base}.{sub_index}"
            index_cell.value = new_val
            count += 1

print(f"Updated {count} sub-indices in sheet '{sheet_name}'.")
wb.save(file_path)
print("Finished.")

