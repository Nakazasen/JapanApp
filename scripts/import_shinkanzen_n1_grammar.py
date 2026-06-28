import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))


import openpyxl
import os
import re
from sqlmodel import Session, create_engine, select
from frontend.core.config import settings
from frontend.models.grammar import GrammarTopic, GrammarCategory, GrammarExample, GrammarMasteryStatus

# --- Configurations ---
project_root = r"C:\ProgramData\Sandbox\Projects\EnglishApp"
file_path = os.path.join(project_root, "[Tailieutiengnhat.net]_tu-vung-tieng-nhat-n1-day-du.xlsx")
sheet_name = "譁ｰ螳悟・譁・ｳ逼1"
SOURCE = "Shin Kanzen Master"
LEVEL = "N1"
USER_ID = 1 # Admin user

def import_grammar():
    database_url = f"sqlite:///{settings.db_path}"
    engine = create_engine(database_url)
    
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    print(f"Opening Excel sheet: {sheet_name}...")
    # Using read_only=True for faster reading if needed, but we need to iterate all anyway
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found.")
        return
        
    ws = wb[sheet_name]
    
    category_map = {} # cache for categories
    
    print("Processing rows...")
    import_count = 0
    
    with Session(engine) as session:
        # We process rows starting from row 1 (headers)
        # Columns: 0:GrpNo, 1:GrpName, 2:ItmNo, 3:Pattern, 4:Connection, 5:Meaning, 6:Notes, 7:Example
        
        current_group_name = "N1 Grammar" # Default
        
        for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
            # row[0] might be group number, row[1] group name
            if row[1] and str(row[1]).strip():
                current_group_name = str(row[1]).strip()
            
            # Find or create category
            if current_group_name not in category_map:
                statement = select(GrammarCategory).where(GrammarCategory.name == current_group_name)
                cat = session.exec(statement).first()
                if not cat:
                    cat = GrammarCategory(
                        name=current_group_name,
                        user_id=USER_ID,
                        lang="jp",
                        icon="燈"
                    )
                    session.add(cat)
                    session.commit()
                    session.refresh(cat)
                category_map[current_group_name] = cat.id
            
            cat_id = category_map[current_group_name]
            
            # Extract grammar data
            pattern_title = row[3]
            connection = row[4]
            meaning_vi = row[5]
            notes = row[6]
            example_block = row[7]
            
            if not pattern_title or pattern_title == "譎る俣髢｢菫・: # Skip headers/metadata
                continue
            if i == 0: continue # Header row if exists (though analysis shows row 1 has data)
            
            # Check if this pattern already exists to avoid duplicates
            statement = select(GrammarTopic).where(
                GrammarTopic.title == str(pattern_title),
                GrammarTopic.source_material == SOURCE
            )
            existing = session.exec(statement).first()
            if existing:
                continue
                
            # Create Grammar Topic
            # We'll put meaning_vi at the top of description
            full_description = f"**ﾃ・nghﾄｩa:** {meaning_vi}\n\n**Gi蘯｣i thﾃｭch:**\n{notes}" if meaning_vi else str(notes)
            
            topic = GrammarTopic(
                user_id=USER_ID,
                lang="jp",
                title=str(pattern_title),
                pattern=str(connection) if connection else "",
                description=full_description,
                usage_notes=str(notes) if notes else "",
                category_id=cat_id,
                level=LEVEL,
                source_material=SOURCE,
                mastery_status=GrammarMasteryStatus.NEW.value
            )
            session.add(topic)
            session.commit()
            session.refresh(topic)
            
            # Parse Example
            if example_block:
                # Format: [Translation]\n[Japanese Example] or vice versa
                # Sample: "Ngﾃy xu蘯･t phﾃ｡t... \n 遨ｺ貂ｯ縺ｫ逹縺上′譌ｩ縺・°"
                parts = str(example_block).split('\n')
                if len(parts) >= 2:
                    vi_trans = parts[0].strip()
                    jp_text = parts[1].strip()
                else:
                    jp_text = parts[0].strip()
                    vi_trans = ""
                    
                example = GrammarExample(
                    topic_id=topic.id,
                    example_text=jp_text,
                    translation_vi=vi_trans
                )
                session.add(example)
            
            import_count += 1
            if import_count % 50 == 0:
                print(f"Imported {import_count} patterns...")
                session.commit()
                
        session.commit()
        
    print(f"Successfully imported {import_count} Shinkanzen N1 grammar patterns.")

if __name__ == "__main__":
    import_grammar()

